"""Redirection validation url."""

from bs4 import BeautifulSoup
import xlsxwriter
import datetime
from openpyxl import load_workbook
from urllib.error import URLError, HTTPError
import threading
import requests

all_urls = []
total_start_time = datetime.datetime.now()

production_url = ''  # Add wp url here
wb = load_workbook(filename='seo_urls.xlsx', read_only=True)
ws = wb['url']

for row in ws.rows:
    temp_list = []
    for cell in row:
        if cell.value:
            temp_list.append(cell.value.strip())
        else:
            temp_list.append('')
    all_urls.append(tuple(temp_list))

print('There are -> ' + str(len(all_urls)) + ' to validate')

error_url_data = []
error_page_data = []


def scrape_urls(frm, to):
    """Scrape url function."""
    global error_url_data
    global error_page_data

    count = frm
    for url_data in all_urls[frm: to]:
        start_time = datetime.datetime.now()
        url = url_data[0]
        slug = url.split(production_url)[1]
        en_url = production_url + slug
        fr_url = production_url + '/fr' + slug

        # English
        try:
            req = requests.get(en_url, headers={
                'User-Agent': 'Mozilla/5.0',
                'Cache-Control': 'no-cache'
            })
            page3 = req.content
            soup3 = BeautifulSoup(page3, "html.parser")
            head = soup3.find('head')
            tilte_ele = head.find('title')
            discription_ele = head.find('meta', {'name': "description"})
            tilte = tilte_ele.text.strip()
            if discription_ele:
                discription = discription_ele.get('content').strip()
            else:
                discription = ''
            match_title = True
            match_disc = True
            if tilte != url_data[1]:
                match_title = False
            if discription != url_data[2]:
                match_disc = False
            if not match_title or not match_disc:
                error_string = 'Error'
                if not match_title:
                    error_string += ' title'
                if not match_disc:
                    error_string += ' discription'
                temp_tuple = (
                    en_url,
                    url_data[1],
                    tilte,
                    url_data[2],
                    discription,
                    error_string
                )
                error_url_data.append(temp_tuple)
            print('EN -> ' + str(count) + '/' + str(len(all_urls)) + ' -> time taken ' + str(datetime.datetime.now() - start_time) +' -> Scraped sub url ' + str(url))
        except HTTPError as e:
            # do something
            print('Fail: Error code: ', e.code, en_url)
            temp_er_dict = dict()
            temp_er_dict['site'] = en_url
            temp_er_dict['error'] = e.code
            error_page_data.append(temp_er_dict)
        except URLError as e:
            # do something (set req to blank)
            print('Fail: Reason: ', e.reason, en_url)
            temp_er_dict = dict()
            temp_er_dict['site'] = en_url
            temp_er_dict['error'] = e.reason
            error_page_data.append(temp_er_dict)
        except Exception as e:
            print('Fail: Reason: ', e, en_url)
            temp_er_dict = dict()
            temp_er_dict['site'] = en_url
            temp_er_dict['error'] = e
            error_page_data.append(temp_er_dict)

        # French
        try:
            req = requests.get(fr_url, headers={'User-Agent': 'Mozilla/5.0', 'Cache-Control': 'no-cache'})
            page3 = req.content
            soup3 = BeautifulSoup(page3, "html.parser")
            head = soup3.find('head')
            tilte_ele = head.find('title')
            discription_ele = head.find('meta', {'name': "description"})
            tilte = tilte_ele.text.strip()
            if discription_ele:
                discription = discription_ele.get('content').strip()
            else:
                discription = ''
            match_title = True
            match_disc = True
            if tilte != url_data[3]:
                match_title = False
            if discription != url_data[4]:
                match_disc = False
            if not match_title or not match_disc:
                error_string = 'Error'
                if not match_title:
                    error_string += ' title'
                if not match_disc:
                    error_string += ' discription'
                temp_tuple = (
                    fr_url,
                    url_data[3],
                    tilte,
                    url_data[4],
                    discription,
                    error_string
                )
                error_url_data.append(temp_tuple)
            print('FR -> ' + str(count) + '/' + str(len(all_urls)) + ' -> time taken ' + str(datetime.datetime.now() - start_time) +' -> Scraped sub url ' + str(fr_url))
        except HTTPError as e:
            # do something
            print('Fail: Error code: ', e.code, fr_url)
            temp_er_dict = dict()
            temp_er_dict['site'] = fr_url
            temp_er_dict['error'] = e.code
            error_page_data.append(temp_er_dict)
        except URLError as e:
            # do something (set req to blank)
            print('Fail: Reason: ', e.reason, fr_url)
            temp_er_dict = dict()
            temp_er_dict['site'] = fr_url
            temp_er_dict['error'] = e.reason
            error_page_data.append(temp_er_dict)
        except Exception as e:
            print('Fail: Reason: ', e, fr_url)
            temp_er_dict = dict()
            temp_er_dict['site'] = fr_url
            temp_er_dict['error'] = e
            error_page_data.append(temp_er_dict)
        count += 1


def run():
    """Run function."""
    t1 = threading.Thread(target=scrape_urls, args=(0, 10))
    t2 = threading.Thread(target=scrape_urls, args=(10, 20))
    t3 = threading.Thread(target=scrape_urls, args=(20, 30))
    t4 = threading.Thread(target=scrape_urls, args=(30, 40))
    t5 = threading.Thread(target=scrape_urls, args=(40, 50))
    t6 = threading.Thread(target=scrape_urls, args=(50, 60))
    t7 = threading.Thread(target=scrape_urls, args=(60, 70))
    t8 = threading.Thread(target=scrape_urls, args=(70, 80))
    t9 = threading.Thread(target=scrape_urls, args=(80, len(all_urls) + 1))
    # t10 = threading.Thread(target=scrape_urls, args=(90, len(all_urls) + 1))

    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    t6.start()
    t7.start()
    t8.start()
    t9.start()
    # t10.start()

    t1.join()
    t2.join()
    t3.join()
    t4.join()
    t5.join()
    t6.join()
    t7.join()
    t8.join()
    t9.join()
    # t10.join()

    try:
        workbook = xlsxwriter.Workbook(
            'seo_error' + str(datetime.datetime.now()) + '.xlsx'
        )
        worksheet_error = workbook.add_worksheet('error_data')
        worksheet2 = workbook.add_worksheet('error_page')

        number1 = 0
        for each in error_url_data:
            worksheet_error.write(
                number1,
                0,
                datetime.datetime.now().strftime('%x')
            )
            worksheet_error.write(number1, 1, each[0])
            worksheet_error.write(number1, 2, each[1])
            worksheet_error.write(number1, 3, each[2])
            worksheet_error.write(number1, 4, each[3])
            worksheet_error.write(number1, 5, each[4])
            worksheet_error.write(number1, 6, each[5])
            number1 += 1

        number2 = 0
        for each in error_page_data:
            worksheet2.write(
                number2, 0, datetime.datetime.now().strftime('%x')
            )
            worksheet2.write(number2, 1, each['site'])
            worksheet2.write(number2, 2, str(each['error']))
            number2 += 1

        workbook.close()
    except Exception as e:
        print('Fail: Reason: ', e)

    print('Completed all total time taken is ' + str(
        datetime.datetime.now() - total_start_time
    ))

run()
