"""Function to Check all inline colours added in the site."""

from bs4 import BeautifulSoup
import xlsxwriter
import datetime
import urllib.request
from openpyxl import load_workbook
from urllib.error import URLError, HTTPError
import threading

all_urls = []

wb = load_workbook(filename='all_urls.xlsx', read_only=True)
ws = wb['all_url']

for row in ws.rows:
    for cell in row:
        all_urls.append(cell.value.strip())

error_data = []
error_page_data = []
all_urls_on_site_data = []


def scrape_urls(frm, to):
    """Scrape URL function passing the from and to will scrape the URL from."""
    global error_data
    global error_page_data
    global all_urls_on_site_data
    count = frm
    for url in all_urls[frm: to]:
        start_time = datetime.datetime.now()
        try:
            page3 = urllib.request.urlopen(url)
            soup3 = BeautifulSoup(page3, "html.parser")
            b_url_list = []
            all_tags_arr = []
            tag_list = ['div', 'span', 'p', 'td', 'tr', 'th', 'ul', 'li', 'a', 'input', 'button', 'label', 'title', 'h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'b', 'em', 'i', 'small', 'ol']
            for taa in tag_list:
                all_tags_arr = all_tags_arr + soup3.find_all(taa)
            for tags in all_tags_arr:
                inline_color = tags.get('color')
                inline_style = tags.get('style')
                if inline_style:
                    aa = inline_style.find('-color:')
                    if inline_style.find('color:') >= 0 and aa == -1:
                        code = inline_style[inline_style.find('color:') + 6:inline_style.find('color:') + 6 + 8].strip().upper()
                        if code not in ['#0072CE', '#4A4A4A', '#F3F5F7', '#333F48', '#FFFFFF', '#000000', '#FFF', '#000']:
                            temp_tuple = (tags, inline_style, tags.text)
                            b_url_list.append(temp_tuple)
                if inline_color and inline_color not in ['#0072CE', '#4A4A4A', '#F3F5F7', '#333F48', '#FFFFFF', '#000000', '#FFF', '#000']:
                    temp_tuple = (tags, inline_color)
                    b_url_list.append(temp_tuple)
            if len(b_url_list) > 0:
                temp_dict = dict()
                temp_dict['site'] = url
                temp_dict['error'] = b_url_list
                error_data.append(temp_dict)
            print(str(count) + '/' + str(len(all_urls)) + ' -> time taken ' + str(datetime.datetime.now() - start_time) +' -> Scraped sub url ' + str(url) + 'found issue ' + str(len(b_url_list)))
        except HTTPError as e:
            # do something
            print('Fail: Error code: ', e.code, url)
            temp_er_dict = dict()
            temp_er_dict['site'] = url
            temp_er_dict['error'] = e.code
            error_page_data.append(temp_er_dict)
        except URLError as e:
            # do something (set req to blank)
            print('Fail: Reason: ', e.reason, url)
            temp_er_dict = dict()
            temp_er_dict['site'] = url
            temp_er_dict['error'] = e.reason
            error_page_data.append(temp_er_dict)
        except Exception as e:
            print('Fail: Reason: ', e, url)
            temp_er_dict = dict()
            temp_er_dict['site'] = url
            temp_er_dict['error'] = e
            error_page_data.append(temp_er_dict)
        count += 1


def run():
    """Run function."""
    total_start_time = datetime.datetime.now()

    t1 = threading.Thread(target=scrape_urls, args=(0, 300))
    t2 = threading.Thread(target=scrape_urls, args=(300, 600))
    t3 = threading.Thread(target=scrape_urls, args=(600, 900))
    t4 = threading.Thread(target=scrape_urls, args=(900, 1200))
    t5 = threading.Thread(target=scrape_urls, args=(1200, 1500))
    t6 = threading.Thread(target=scrape_urls, args=(1500, 1800))
    t7 = threading.Thread(target=scrape_urls, args=(1800, 2100))
    t8 = threading.Thread(target=scrape_urls, args=(2100, 2400))
    t9 = threading.Thread(target=scrape_urls, args=(2400, 2700))
    t10 = threading.Thread(target=scrape_urls, args=(2700, len(all_urls) + 1))

    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    t6.start()
    t7.start()
    t8.start()
    t9.start()
    t10.start()

    t1.join()
    t2.join()
    t3.join()
    t4.join()
    t5.join()
    t6.join()
    t7.join()
    t8.join()
    t9.join()
    t10.join()

    try:
        workbook = xlsxwriter.Workbook('error_color_' + str(datetime.datetime.now()) + '.xlsx')
        worksheet1 = workbook.add_worksheet('error_url')
        worksheet2 = workbook.add_worksheet('error_pages')

        number1 = 0
        for each in error_data:
            worksheet1.write(number1, 0, datetime.datetime.now().strftime('%x'))
            worksheet1.write(number1, 1, each['site'])

            for each1 in each['error']:
                worksheet1.write(number1, 2, str(each1[0]))
                worksheet1.write(number1, 3, str(each1[1]))
                worksheet1.write(number1, 5, str(each1[2]))
                number1 += 1

        number2 = 0
        for each in error_page_data:
            worksheet2.write(number2, 0, datetime.datetime.now().strftime('%x'))
            worksheet2.write(number2, 1, str(each['site']))
            worksheet2.write(number2, 2, str(each['error']))
            number2 += 1

        workbook.close()

    except Exception as e:
        print('Exception occurred ', e)
    print('Completed all total time taken is ' + str(datetime.datetime.now() - total_start_time))

run()
