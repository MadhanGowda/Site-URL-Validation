"""Validate all urls."""

from bs4 import BeautifulSoup
import xlsxwriter
import datetime
from urllib.request import Request, urlopen
from openpyxl import load_workbook
from urllib.error import URLError, HTTPError
import threading

all_urls = []
total_start_time = datetime.datetime.now()

production_url = 'https://www.cppinvestments.com/'
wb = load_workbook(filename='all_links.xlsx', read_only=True)
ws = wb['all_link']

for row in ws.rows:
    temp_list = []
    for cell in row:
        if cell.value:
            temp_list.append(cell.value.strip())
        else:
            temp_list.append('')
    all_urls.append(tuple(temp_list))

all_links = []
for each in all_urls:
    all_links.append(each[2])

all_links = list(set(all_links))
print('There are -> ' + str(len(all_links)) + ' to validate')


error_url_data = []

exclude_urls = []


def scrape_urls(frm, to):
    """Scrape url function."""
    global error_url_data

    count = frm
    for url1 in all_links[frm: to]:
        start_time = datetime.datetime.now()
        url = url1

        if url not in exclude_urls and url[:7] != 'mailto:' and url[:4] != 'tel:' and not (url[0] == '#' and len(url) > 1):
            if url[0] == '/':
                url = production_url + url
            try:
                req = Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                page3 = urlopen(req)
                soup3 = BeautifulSoup(page3, "html.parser")
                print(str(count) + '/' + str(len(all_links)) + ' -> time taken ' + str(datetime.datetime.now() - start_time) +' -> Scraped sub url ' + str(url))
            except HTTPError as e:
                # do something
                print('Fail: Error code: ', e.code, url)
                temp_er_dict = dict()
                temp_er_dict['site'] = url
                temp_er_dict['error'] = e.code
                error_url_data.append(temp_er_dict)
            except URLError as e:
                # do something (set req to blank)
                print('Fail: Reason: ', e.reason, url)
                temp_er_dict = dict()
                temp_er_dict['site'] = url
                temp_er_dict['error'] = e.reason
                error_url_data.append(temp_er_dict)
            except Exception as e:
                print('Fail: Reason: ', e, url)
                temp_er_dict = dict()
                temp_er_dict['site'] = url
                temp_er_dict['error'] = e
                error_url_data.append(temp_er_dict)
            count += 1


def find_all_in_urls_in_list(link):
    """Find all the urls in the list."""
    error_list = []
    for link_url in all_urls:
        if link_url[2] == link:
            error_list.append(link_url)
    return error_list


def run():
    """Run script."""
    t1 = threading.Thread(target=scrape_urls, args=(0, 1000))
    t2 = threading.Thread(target=scrape_urls, args=(1000, 2000))
    t3 = threading.Thread(target=scrape_urls, args=(2000, 3000))
    t4 = threading.Thread(target=scrape_urls, args=(3000, 4000))
    t5 = threading.Thread(target=scrape_urls, args=(4000, 5000))
    t6 = threading.Thread(target=scrape_urls, args=(5000, 6000))
    t7 = threading.Thread(target=scrape_urls, args=(6000, 7000))
    t8 = threading.Thread(target=scrape_urls, args=(7000, 8000))
    t9 = threading.Thread(target=scrape_urls, args=(8000, 9000))
    t10 = threading.Thread(target=scrape_urls, args=(9000, 10000))
    t11 = threading.Thread(target=scrape_urls, args=(10000, 11000))
    t12 = threading.Thread(target=scrape_urls, args=(11000, 12000))
    t13 = threading.Thread(target=scrape_urls, args=(12000, 13000))
    t14 = threading.Thread(target=scrape_urls, args=(13000, len(all_links) + 1))

    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    t6.start()
    t7.start()
    t8.start()
    t9.start()
    t10.start()
    t11.start()
    t12.start()
    t13.start()
    t14.start()

    t1.join()
    t2.join()
    t3.join()
    t4.join()
    t5.join()
    t6.join()
    t7.join()
    t8.join()
    t9.join()
    t10.join()
    t11.join()
    t12.join()
    t13.join()
    t14.join()

    try:
        workbook = xlsxwriter.Workbook('all_error_url' + str(datetime.datetime.now()) + '.xlsx')
        worksheet_error = workbook.add_worksheet('error_pages')
        worksheet_hash = workbook.add_worksheet('hash')
        worksheet_hash_id = workbook.add_worksheet('hash_id')
        worksheet_empty = workbook.add_worksheet('empty')
        worksheet_void = workbook.add_worksheet('void')
        worksheet_mail = workbook.add_worksheet('mail')
        worksheet_tel = workbook.add_worksheet('tele')

        number1 = 0
        for each in error_url_data:
            all_posible_places = find_all_in_urls_in_list(each['site'])
            for each1 in all_posible_places:
                worksheet_error.write(number1, 0, datetime.datetime.now().strftime('%x'))
                worksheet_error.write(number1, 1, each1[1])
                worksheet_error.write(number1, 2, each1[2])
                worksheet_error.write(number1, 3, each1[3])
                worksheet_error.write(number1, 4, str(each['error']))
                number1 += 1

        all_posible_places_hash = find_all_in_urls_in_list('#')
        number2 = 0
        for each1 in all_posible_places_hash:
            worksheet_hash.write(number2, 0, datetime.datetime.now().strftime('%x'))
            worksheet_hash.write(number2, 1, each1[1])
            worksheet_hash.write(number2, 2, each1[2])
            worksheet_hash.write(number2, 3, each1[3])
            number2 += 1

        all_posible_places_empty = find_all_in_urls_in_list('')
        number3 = 0
        for each1 in all_posible_places_empty:
            worksheet_empty.write(number3, 0, datetime.datetime.now().strftime('%x'))
            worksheet_empty.write(number3, 1, each1[1])
            worksheet_empty.write(number3, 2, each1[2])
            worksheet_empty.write(number3, 3, each1[3])
            number3 += 1

        all_posible_places_void = find_all_in_urls_in_list('javascript:void(0);')
        number4 = 0
        for each1 in all_posible_places_void:
            worksheet_void.write(number4, 0, datetime.datetime.now().strftime('%x'))
            worksheet_void.write(number4, 1, each1[1])
            worksheet_void.write(number4, 2, each1[2])
            worksheet_void.write(number4, 3, each1[3])
            number4 += 1

        all_posible_places_mail = []
        for link_url in all_urls:
            if link_url[2] != '' and link_url[2][:7] == 'mailto:':
                all_posible_places_mail.append(link_url)
        number5 = 0
        for each1 in all_posible_places_mail:
            worksheet_mail.write(number5, 0, datetime.datetime.now().strftime('%x'))
            worksheet_mail.write(number5, 1, each1[1])
            worksheet_mail.write(number5, 2, each1[2])
            worksheet_mail.write(number5, 3, each1[3])
            number5 += 1

        all_posible_places_tel = []
        for link_url in all_urls:
            if link_url[2] != '' and link_url[2][:4] == 'tel:':
                all_posible_places_tel.append(link_url)
        number6 = 0
        for each1 in all_posible_places_tel:
            worksheet_tel.write(number6, 0, datetime.datetime.now().strftime('%x'))
            worksheet_tel.write(number6, 1, each1[1])
            worksheet_tel.write(number6, 2, each1[2])
            worksheet_tel.write(number6, 3, each1[3])
            number6 += 1

        all_posible_places_hash_id = []
        for link_url in all_urls:
            if link_url[2] != '' and link_url[2][0] == '#' and len(link_url[2]) > 1:
                all_posible_places_hash_id.append(link_url)
        number7 = 0
        for each1 in all_posible_places_hash_id:
            worksheet_hash_id.write(number7, 0, datetime.datetime.now().strftime('%x'))
            worksheet_hash_id.write(number7, 1, each1[1])
            worksheet_hash_id.write(number7, 2, each1[2])
            worksheet_hash_id.write(number7, 3, each1[3])
            number7 += 1

        workbook.close()
    except Exception as e:
        print('Fail: Reason: ', e)

    print('Completed all total time taken is ' + str(datetime.datetime.now() - total_start_time))

run()
