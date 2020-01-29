"""Function to scrape all the URLs and writes an Excel."""

from bs4 import BeautifulSoup
import xlsxwriter
import datetime
import urllib.request
from openpyxl import load_workbook
from urllib.error import URLError, HTTPError
import threading

bad_urls = []  # Add bad URLs that needs to be errored.
good_urls = []  # Add good URLs that needs to be skipped.

all_urls = []
wb = load_workbook(filename='all_urls.xlsx', read_only=True)
ws = wb['all_url']

for row in ws.rows:
    for cell in row:
        all_urls.append(cell.value.strip())

error_data = []
error_image = []
error_page_data = []
all_urls_on_site_data = []
all_imgs_on_site_data = []


def scrape_urls(frm, to):
    """Scrape URL function passing the from and to will scrape the URL from."""
    global error_data
    global error_image
    global error_page_data
    global all_urls_on_site_data
    global all_imgs_on_site_data
    count = frm
    for url in all_urls[frm: to]:
        start_time = datetime.datetime.now()
        try:
            page3 = urllib.request.urlopen(url)
            soup3 = BeautifulSoup(page3, "html.parser")
            b_url_list = []
            b_img_list = []
            for link in soup3.findAll('a'):
                val_url = link.get('href')
                link_text = link.text
                bad_url = False
                good_url = False
                for g_url in good_urls:
                    if val_url and val_url.find(g_url) > 0:
                        good_url = True
                if not good_url:
                    for b_url in bad_urls:
                        if val_url and val_url.find(b_url) > 0:
                            temp_tuple = (val_url, link_text)
                            b_url_list.append(temp_tuple)
                            bad_url = True
                if not bad_url:
                    temp_tuple2 = (url, val_url, link_text)
                    all_urls_on_site_data.append(temp_tuple2)
            for img in soup3.findAll('img'):
                val_src = img.get('src')
                link_text = img.get('alt')
                bad_url = False
                good_url = False
                for g_url in good_urls:
                    if val_src and val_src.find(g_url) > 0:
                        good_url = True
                if not good_url:
                    for b_url in bad_urls:
                        if val_src and val_src.find(b_url) > 0:
                            temp_tuple = (val_src, link_text)
                            b_img_list.append(temp_tuple)
                            bad_url = True
                if not bad_url:
                    temp_tuple2 = (url, val_src, link_text)
                    all_imgs_on_site_data.append(temp_tuple2)
            if len(b_url_list) > 0:
                temp_dict = dict()
                temp_dict['site'] = url
                temp_dict['error'] = b_url_list
                error_data.append(temp_dict)
            if len(b_img_list) > 0:
                temp_dict = dict()
                temp_dict['site'] = url
                temp_dict['error'] = b_img_list
                error_image.append(temp_dict)
            print(
                str(count) + '/' + str(len(all_urls)) + ' -> time taken ' + str(datetime.datetime.now() - start_time) +' -> Scraped sub url ' + str(url) + 'found issue ' + str(len(b_url_list)), b_url_list)
        except HTTPError as e:
            # do something
            print('Fail: Error code: ', e.code, url)
            temp_er_dict = dict()
            temp_er_dict['site'] = url
            temp_er_dict['error'] = e.code
            error_page_data.append(temp_er_dict)
        except URLError as e:
            # do something (set req to blank)
            print('Fail: Reason: ', e.reason, url)
            temp_er_dict = dict()
            temp_er_dict['site'] = url
            temp_er_dict['error'] = e.reason
            error_page_data.append(temp_er_dict)
        except Exception as e:
            print('Fail: Reason: ', e, url)
            temp_er_dict = dict()
            temp_er_dict['site'] = url
            temp_er_dict['error'] = e
            error_page_data.append(temp_er_dict)
        count += 1


def run():

    total_start_time = datetime.datetime.now()

    t1 = threading.Thread(target=scrape_urls, args=(0, 300))
    t2 = threading.Thread(target=scrape_urls, args=(300, 600))
    t3 = threading.Thread(target=scrape_urls, args=(600, 900))
    t4 = threading.Thread(target=scrape_urls, args=(900, 1200))
    t5 = threading.Thread(target=scrape_urls, args=(1200, 1500))
    t6 = threading.Thread(target=scrape_urls, args=(1500, 1800))
    t7 = threading.Thread(target=scrape_urls, args=(1800, 2100))
    t8 = threading.Thread(target=scrape_urls, args=(2100, 2400))
    t9 = threading.Thread(target=scrape_urls, args=(2400, 2700))
    t10 = threading.Thread(target=scrape_urls, args=(2700, len(all_urls) + 1))

    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t5.start()
    t6.start()
    t7.start()
    t8.start()
    t9.start()
    t10.start()

    t1.join()
    t2.join()
    t3.join()
    t4.join()
    t5.join()
    t6.join()
    t7.join()
    t8.join()
    t9.join()
    t10.join()

    try:
        workbook = xlsxwriter.Workbook('error_url_' + str(datetime.datetime.now()) + '.xlsx')
        worksheet1 = workbook.add_worksheet('error_url')
        worksheet2 = workbook.add_worksheet('error_pages')
        worksheet3 = workbook.add_worksheet('error_image')

        workbook2 = xlsxwriter.Workbook('all_links.xlsx', {'strings_to_urls': False})
        worksheet21 = workbook2.add_worksheet('all_link')

        number1 = 0
        for each in error_data:
            worksheet1.write(number1, 0, datetime.datetime.now().strftime('%x'))
            worksheet1.write(number1, 1, each['site'])

            for each1 in each['error']:
                worksheet1.write(number1, 2, each1[0])
                worksheet1.write(number1, 3, each1[1])
                number1 += 1

        number4 = 0
        for each in error_image:
            worksheet3.write(number4, 0, datetime.datetime.now().strftime('%x'))
            worksheet3.write(number4, 1, each['site'])

            for each1 in each['error']:
                worksheet3.write(number4, 2, each1[0])
                worksheet3.write(number4, 3, each1[1])
                number4 += 1

        number2 = 0
        for each in error_page_data:
            worksheet2.write(number2, 0, datetime.datetime.now().strftime('%x'))
            worksheet2.write(number2, 1, each['site'])
            worksheet2.write(number2, 2, each['error'])
            number2 += 1

        number3 = 0
        unique_urls = list(set(all_urls_on_site_data))
        for each in unique_urls:
            worksheet21.write(number3, 0, datetime.datetime.now().strftime('%x'))
            worksheet21.write(number3, 1, each[0])
            worksheet21.write(number3, 2, each[1])
            worksheet21.write(number3, 3, each[2])
            number3 += 1

        workbook.close()
        workbook2.close()

    except Exception as e:
        print('Exception occurred ', e)
    print('Completed all total time taken is ' + str(datetime.datetime.now() - total_start_time))

run()
